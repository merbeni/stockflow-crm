"""
Genera el documento de Casos de Prueba de StockFlow CRM en formato Excel (.xlsx).
Basado en la suite de tests pytest (117 tests backend).
Ejecutar: python generate_test_cases.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── palette ────────────────────────────────────────────────────────────────────
C_HEADER  = "1A56DB"
C_PASS    = "D1FAE5"
C_FAIL    = "FEE2E2"
C_PEND    = "FEF9C3"
C_BORDER  = "9CA3AF"
C_ROW_A   = "F9FAFB"
C_ROW_B   = "FFFFFF"

MODULE_COLORS = {
    "Autenticación":       "DBEAFE",
    "Seguridad":           "EDE9FE",
    "Productos":           "FEF9C3",
    "Proveedores":         "FCE7F3",
    "Facturas (IA)":       "E0F2FE",
    "Clientes":            "FEE2E2",
    "Pedidos":             "D1FAE5",
    "Movimientos de Stock":"F3F4F6",
}


def _side(c=C_BORDER): return Side(style="thin", color=c)
def _border(): return Border(left=_side(), right=_side(), top=_side(), bottom=_side())
def _fill(h): return PatternFill("solid", fgColor=h)
def _font(sz=9, bold=False, color="000000"): return Font(name="Calibri", size=sz, bold=bold, color=color)
def _wrap(h="left"): return Alignment(wrap_text=True, vertical="top", horizontal=h)


# ── test case data ─────────────────────────────────────────────────────────────
# Columns: id, module, name, preconditions, input_action, expected_result, status
TEST_CASES = [
    # ═══ AUTENTICACIÓN ════════════════════════════════════════════════════════
    ("CP-001","Autenticación","Registro exitoso de usuario con rol operator",
     "DB vacía","POST /auth/register {email, password, role:'operator'}",
     "HTTP 201 — devuelve id, email, role. No expone hashed_password.","Pasa"),
    ("CP-002","Autenticación","Registro con rol admin",
     "DB vacía","POST /auth/register {email, password, role:'admin'}",
     "HTTP 201 — role='admin'.","Pasa"),
    ("CP-003","Autenticación","Rol por defecto es operator cuando no se envía",
     "DB vacía","POST /auth/register {email, password} (sin role)",
     "HTTP 201 — role='operator'.","Pasa"),
    ("CP-004","Autenticación","Registro con email duplicado devuelve 400",
     "Email ya registrado","POST /auth/register con el mismo email",
     "HTTP 400 — detail contiene 'already registered'.","Pasa"),
    ("CP-005","Autenticación","Email con formato inválido devuelve 422",
     "DB vacía","POST /auth/register {email:'not-an-email', password:'...'}",
     "HTTP 422 — error de validación Pydantic.","Pasa"),
    ("CP-006","Autenticación","Registro sin contraseña devuelve 422",
     "DB vacía","POST /auth/register {email:'user@test.com'} (sin password)",
     "HTTP 422.","Pasa"),
    ("CP-007","Autenticación","Login exitoso devuelve token JWT",
     "Usuario registrado","POST /auth/login {email, password correctos}",
     "HTTP 200 — devuelve access_token y token_type='bearer'.","Pasa"),
    ("CP-008","Autenticación","Login con contraseña incorrecta devuelve 401",
     "Usuario registrado","POST /auth/login {email correcto, password incorrecto}",
     "HTTP 401.","Pasa"),
    ("CP-009","Autenticación","Login con email desconocido devuelve 401",
     "DB vacía","POST /auth/login {email no registrado, password cualquiera}",
     "HTTP 401.","Pasa"),
    ("CP-010","Autenticación","GET /auth/me devuelve usuario actual",
     "Usuario autenticado","GET /auth/me con token válido",
     "HTTP 200 — devuelve email='admin@test.com'.","Pasa"),
    ("CP-011","Autenticación","GET /auth/me sin token devuelve 401",
     "Sin token","GET /auth/me sin header Authorization",
     "HTTP 401.","Pasa"),
    ("CP-012","Autenticación","GET /auth/me con token inválido devuelve 401",
     "Token malformado","GET /auth/me con 'Bearer invalid.token'",
     "HTTP 401.","Pasa"),

    # ═══ SEGURIDAD ════════════════════════════════════════════════════════════
    ("CP-013","Seguridad","Hash de contraseña con bcrypt",
     "—","hash_password('secret123')",
     "El resultado empieza con '$2b$'. Verificable con verify_password.","Pasa"),
    ("CP-014","Seguridad","Verificación correcta de contraseña hasheada",
     "Password hasheado","verify_password('secret123', hash)",
     "Retorna True.","Pasa"),
    ("CP-015","Seguridad","Verificación incorrecta de contraseña",
     "Password hasheado","verify_password('wrongpass', hash)",
     "Retorna False.","Pasa"),
    ("CP-016","Seguridad","Creación de token JWT contiene sub y exp",
     "—","create_access_token({'sub': 'test@test.com'})",
     "JWT decodificable que contiene 'sub' y 'exp'.","Pasa"),
    ("CP-017","Seguridad","Token JWT no expirado se valida correctamente",
     "Token válido","decode_access_token(token válido)",
     "Devuelve payload con sub='test@test.com'.","Pasa"),
    ("CP-018","Seguridad","Token JWT expirado lanza JWTError",
     "Token con exp en el pasado","decode_access_token(token expirado)",
     "Lanza jose.JWTError.","Pasa"),
    ("CP-019","Seguridad","Token firmado con clave incorrecta es rechazado",
     "Token firmado con otra clave","decode_access_token(token con firma inválida)",
     "Lanza jose.JWTError.","Pasa"),
    ("CP-020","Seguridad","Payload malformado lanza JWTError",
     "—","decode_access_token('not.a.jwt')",
     "Lanza jose.JWTError.","Pasa"),
    ("CP-021","Seguridad","create_access_token acepta timedelta personalizado",
     "—","create_access_token({'sub':'u'}, expires_delta=timedelta(minutes=5))",
     "Token creado; exp ≈ now + 5 min.","Pasa"),

    # ═══ PRODUCTOS ════════════════════════════════════════════════════════════
    ("CP-022","Productos","Crear producto exitosamente",
     "DB vacía, usuario autenticado","POST /products {sku, name, price, current_stock, minimum_stock}",
     "HTTP 201 — devuelve producto con is_active=True.","Pasa"),
    ("CP-023","Productos","Crear producto con stock > 0 genera movimiento de entrada",
     "DB vacía","POST /products {current_stock:'25.000'}",
     "Se crea un StockMovement de tipo 'entry' con quantity=25.","Pasa"),
    ("CP-024","Productos","Crear producto con stock = 0 no genera movimiento",
     "DB vacía","POST /products {current_stock:'0.000'}",
     "No existe ningún StockMovement para ese producto.","Pasa"),
    ("CP-025","Productos","SKU duplicado devuelve 400",
     "Producto 'DUP-SKU' ya existe","POST /products {sku:'DUP-SKU', ...}",
     "HTTP 400 — detail contiene 'already exists'.","Pasa"),
    ("CP-026","Productos","Crear producto sin autenticación devuelve 401",
     "Sin token","POST /products sin header Authorization",
     "HTTP 401.","Pasa"),
    ("CP-027","Productos","Listar productos — lista vacía",
     "Sin productos en DB","GET /products",
     "HTTP 200 — devuelve [].","Pasa"),
    ("CP-028","Productos","Listar productos devuelve todos",
     "2 productos en DB","GET /products",
     "HTTP 200 — array de 2 elementos.","Pasa"),
    ("CP-029","Productos","Filtro low_stock_only filtra correctamente",
     "Producto Low (stock<min) y OK (stock≥min)","GET /products?low_stock_only=true",
     "Sólo devuelve producto 'Low'.","Pasa"),
    ("CP-030","Productos","Campo low_stock calculado correctamente",
     "Producto con stock=3 y min=10","GET /products",
     "low_stock=True en ese producto.","Pasa"),
    ("CP-031","Productos","Obtener producto existente",
     "Producto creado","GET /products/{id}",
     "HTTP 200 — id coincide.","Pasa"),
    ("CP-032","Productos","Obtener producto inexistente devuelve 404",
     "DB vacía","GET /products/9999",
     "HTTP 404.","Pasa"),
    ("CP-033","Productos","Actualizar nombre y precio",
     "Producto existente","PUT /products/{id} {name:'Updated', price:'19.99'}",
     "HTTP 200 — name y price actualizados.","Pasa"),
    ("CP-034","Productos","Actualizar stock crea movimiento de ajuste positivo",
     "Producto con stock=50","PUT /products/{id} {current_stock:'60.000'}",
     "Movimiento 'adjustment' con quantity=10.","Pasa"),
    ("CP-035","Productos","Actualizar stock crea movimiento de ajuste negativo",
     "Producto con stock=50","PUT /products/{id} {current_stock:'30.000'}",
     "Movimiento 'adjustment' con quantity=-20.","Pasa"),
    ("CP-036","Productos","Actualizar producto inexistente devuelve 404",
     "DB vacía","PUT /products/9999 {name:'Ghost'}",
     "HTTP 404.","Pasa"),
    ("CP-037","Productos","Eliminar producto sin movimientos exitosamente",
     "Producto sin historial (stock=0)","DELETE /products/{id}",
     "HTTP 204 — GET posterior devuelve 404.","Pasa"),
    ("CP-038","Productos","Eliminar producto con movimientos devuelve 409",
     "Producto con stock>0 (tiene movimiento de entrada)","DELETE /products/{id}",
     "HTTP 409.","Pasa"),
    ("CP-039","Productos","Eliminar producto inexistente devuelve 404",
     "DB vacía","DELETE /products/9999",
     "HTTP 404.","Pasa"),

    # ═══ PROVEEDORES ══════════════════════════════════════════════════════════
    ("CP-040","Proveedores","Crear proveedor con todos los campos",
     "DB vacía","POST /suppliers {name, contact_name, email, phone}",
     "HTTP 201 — devuelve todos los campos incluyendo id.","Pasa"),
    ("CP-041","Proveedores","Crear proveedor sin teléfono (campo opcional)",
     "DB vacía","POST /suppliers {name, contact_name, email} (sin phone)",
     "HTTP 201 — phone=null.","Pasa"),
    ("CP-042","Proveedores","Crear proveedor sin campos obligatorios devuelve 422",
     "DB vacía","POST /suppliers {name:'Incomplete'} (sin contact_name ni email)",
     "HTTP 422.","Pasa"),
    ("CP-043","Proveedores","Crear proveedor sin autenticación devuelve 401",
     "Sin token","POST /suppliers {name:'Anon'}",
     "HTTP 401.","Pasa"),
    ("CP-044","Proveedores","Listar proveedores — lista vacía",
     "Sin proveedores","GET /suppliers",
     "HTTP 200 — devuelve [].","Pasa"),
    ("CP-045","Proveedores","Listar proveedores devuelve todos",
     "2 proveedores en DB","GET /suppliers",
     "HTTP 200 — array de 2 elementos.","Pasa"),
    ("CP-046","Proveedores","Proveedores se devuelven ordenados por nombre",
     "Proveedores 'Zeta' y 'Alpha'","GET /suppliers",
     "El primer elemento es 'Alpha'.","Pasa"),
    ("CP-047","Proveedores","Obtener proveedor existente",
     "Proveedor creado","GET /suppliers/{id}",
     "HTTP 200 — id coincide.","Pasa"),
    ("CP-048","Proveedores","Obtener proveedor inexistente devuelve 404",
     "DB vacía","GET /suppliers/9999",
     "HTTP 404.","Pasa"),
    ("CP-049","Proveedores","Actualizar nombre del proveedor",
     "Proveedor existente","PUT /suppliers/{id} {name:'New Name'}",
     "HTTP 200 — name='New Name'.","Pasa"),
    ("CP-050","Proveedores","Actualizar proveedor inexistente devuelve 404",
     "DB vacía","PUT /suppliers/9999 {name:'Ghost'}",
     "HTTP 404.","Pasa"),
    ("CP-051","Proveedores","Eliminar proveedor exitosamente",
     "Proveedor existente","DELETE /suppliers/{id}",
     "HTTP 204 — GET posterior devuelve 404.","Pasa"),
    ("CP-052","Proveedores","Eliminar proveedor inexistente devuelve 404",
     "DB vacía","DELETE /suppliers/9999",
     "HTTP 404.","Pasa"),

    # ═══ FACTURAS (IA) ════════════════════════════════════════════════════════
    ("CP-053","Facturas (IA)","Procesar factura exitosamente (Gemini mockeado)",
     "Usuario autenticado. Gemini mock devuelve 2 ítems.",
     "POST /invoices/process {file: invoice.pdf}",
     "HTTP 201 — devuelve invoice_id, 2 items, supplier='Acme Corp', date.","Pasa"),
    ("CP-054","Facturas (IA)","Proceso asocia proveedor si el nombre coincide",
     "Proveedor 'Acme Corp' en DB",
     "POST /invoices/process (Gemini devuelve supplier:'Acme Corp')",
     "supplier_id no es null en la respuesta.","Pasa"),
    ("CP-055","Facturas (IA)","Proceso sugiere producto existente que coincide",
     "Producto 'Blue Widget' en DB",
     "POST /invoices/process (Gemini devuelve ítem 'Blue Widget')",
     "suggested_product_id no es null; suggested_product_name='Blue Widget'.","Pasa"),
    ("CP-056","Facturas (IA)","Sin productos en DB no hay sugerencias",
     "DB sin productos",
     "POST /invoices/process",
     "Todos los ítems tienen suggested_product_id=null.","Pasa"),
    ("CP-057","Facturas (IA)","Proceso requiere autenticación",
     "Sin token",
     "POST /invoices/process sin Authorization",
     "HTTP 401.","Pasa"),
    ("CP-058","Facturas (IA)","Tipo MIME inválido devuelve 415",
     "Usuario autenticado",
     "POST /invoices/process {file: doc.txt, MIME: text/plain}",
     "HTTP 415.","Pasa"),
    ("CP-059","Facturas (IA)","Error de Gemini (JSON inválido) devuelve 422",
     "Gemini mock lanza ValueError",
     "POST /invoices/process {file: invoice.pdf}",
     "HTTP 422.","Pasa"),
    ("CP-060","Facturas (IA)","Confirmar factura con producto existente",
     "Factura pendiente, producto 'BW-001' en DB",
     "POST /invoices/{id}/confirm {items:[{invoice_item_id, product_id}]}",
     "HTTP 200 — status='confirmed'.","Pasa"),
    ("CP-061","Facturas (IA)","Confirmar actualiza el stock del producto",
     "Producto con stock=20, factura con ítem qty=10",
     "POST /invoices/{id}/confirm",
     "current_stock del producto pasa a 30.","Pasa"),
    ("CP-062","Facturas (IA)","Confirmar puede crear producto nuevo inline",
     "Factura pendiente, sin productos en DB",
     "POST /invoices/{id}/confirm {items:[{..., new_product:{sku, name, price}}]}",
     "HTTP 200 — el nuevo producto aparece en GET /products.","Pasa"),
    ("CP-063","Facturas (IA)","Confirmar guarda mapeo SKU proveedor",
     "Factura, proveedor y producto en DB",
     "POST /invoices/{id}/confirm {supplier_id, items:[{..., supplier_sku:'ACME-BW-XYZ'}]}",
     "ProductSupplierMapping creado con supplier_sku='ACME-BW-XYZ'.","Pasa"),
    ("CP-064","Facturas (IA)","Confirmar factura ya confirmada devuelve 400",
     "Factura en estado confirmed",
     "POST /invoices/{id}/confirm nuevamente",
     "HTTP 400.","Pasa"),
    ("CP-065","Facturas (IA)","Rechazar factura pendiente exitosamente",
     "Factura en estado pending",
     "POST /invoices/{id}/reject",
     "HTTP 200 — status='rejected'.","Pasa"),
    ("CP-066","Facturas (IA)","Rechazar factura ya rechazada devuelve 400",
     "Factura en estado rejected",
     "POST /invoices/{id}/reject nuevamente",
     "HTTP 400.","Pasa"),
    ("CP-067","Facturas (IA)","Rechazar factura inexistente devuelve 400",
     "DB vacía",
     "POST /invoices/9999/reject",
     "HTTP 400 — detail contiene 'not found'.","Pasa"),
    ("CP-068","Facturas (IA)","Listar facturas — lista vacía",
     "Sin facturas",
     "GET /invoices",
     "HTTP 200 — devuelve [].","Pasa"),
    ("CP-069","Facturas (IA)","Listar facturas devuelve la factura procesada",
     "1 factura procesada",
     "GET /invoices",
     "Array de longitud 1.","Pasa"),
    ("CP-070","Facturas (IA)","Obtener factura por ID",
     "1 factura procesada",
     "GET /invoices/{id}",
     "HTTP 200 — id coincide, 2 items.","Pasa"),
    ("CP-071","Facturas (IA)","Obtener factura inexistente devuelve 404",
     "DB vacía",
     "GET /invoices/9999",
     "HTTP 404.","Pasa"),

    # ═══ CLIENTES ════════════════════════════════════════════════════════════
    ("CP-072","Clientes","Crear cliente con todos los campos",
     "DB vacía","POST /customers {name, email, phone, address}",
     "HTTP 201 — devuelve todos los campos.","Pasa"),
    ("CP-073","Clientes","Crear cliente sin dirección (opcional)",
     "DB vacía","POST /customers {name, email, phone}",
     "HTTP 201 — address=null.","Pasa"),
    ("CP-074","Clientes","Crear cliente sin autenticación devuelve 401",
     "Sin token","POST /customers {name, email, phone}",
     "HTTP 401.","Pasa"),
    ("CP-075","Clientes","Crear cliente con email inválido devuelve 422",
     "DB vacía","POST /customers {name, email:'not-an-email', phone}",
     "HTTP 422.","Pasa"),
    ("CP-076","Clientes","Listar clientes — lista vacía",
     "Sin clientes","GET /customers",
     "HTTP 200 — devuelve [].","Pasa"),
    ("CP-077","Clientes","Listar clientes devuelve todos",
     "2 clientes en DB","GET /customers",
     "Array de longitud 2.","Pasa"),
    ("CP-078","Clientes","Clientes se devuelven ordenados por nombre",
     "Clientes 'Zoe' y 'Anna'","GET /customers",
     "El primer elemento es 'Anna'.","Pasa"),
    ("CP-079","Clientes","Obtener cliente existente",
     "Cliente creado","GET /customers/{id}",
     "HTTP 200 — id coincide.","Pasa"),
    ("CP-080","Clientes","Obtener cliente inexistente devuelve 404",
     "DB vacía","GET /customers/9999",
     "HTTP 404.","Pasa"),
    ("CP-081","Clientes","Actualizar nombre y teléfono del cliente",
     "Cliente existente","PUT /customers/{id} {name:'Updated', phone:'555-9999'}",
     "HTTP 200 — campos actualizados.","Pasa"),
    ("CP-082","Clientes","Actualizar cliente inexistente devuelve 404",
     "DB vacía","PUT /customers/9999 {name:'Ghost'}",
     "HTTP 404.","Pasa"),
    ("CP-083","Clientes","Eliminar cliente exitosamente",
     "Cliente sin pedidos","DELETE /customers/{id}",
     "HTTP 204 — GET posterior devuelve 404.","Pasa"),
    ("CP-084","Clientes","Eliminar cliente inexistente devuelve 404",
     "DB vacía","DELETE /customers/9999",
     "HTTP 404.","Pasa"),
    ("CP-085","Clientes","Historial de pedidos vacío",
     "Cliente sin pedidos","GET /customers/{id}/orders",
     "HTTP 200 — {customer:{id:...}, orders:[]}.","Pasa"),
    ("CP-086","Clientes","Historial incluye pedidos con total correcto",
     "Cliente con 1 pedido de 2 ítems × $10","GET /customers/{id}/orders",
     "orders tiene 1 elemento con total=20.0.","Pasa"),

    # ═══ PEDIDOS ══════════════════════════════════════════════════════════════
    ("CP-087","Pedidos","Crear pedido exitosamente",
     "Cliente en DB","POST /orders {customer_id}",
     "HTTP 201 — status='pending', items=[], total=0.0.","Pasa"),
    ("CP-088","Pedidos","Crear pedido con cliente inexistente devuelve 400",
     "DB vacía","POST /orders {customer_id:9999}",
     "HTTP 400 — detail contiene 'not found'.","Pasa"),
    ("CP-089","Pedidos","Crear pedido sin autenticación devuelve 401",
     "Sin token","POST /orders {customer_id}",
     "HTTP 401.","Pasa"),
    ("CP-090","Pedidos","Listar pedidos — lista vacía",
     "Sin pedidos","GET /orders",
     "HTTP 200 — devuelve [].","Pasa"),
    ("CP-091","Pedidos","Listar pedidos devuelve el creado",
     "1 pedido en DB","GET /orders",
     "Array de longitud 1.","Pasa"),
    ("CP-092","Pedidos","Obtener pedido existente",
     "Pedido creado","GET /orders/{id}",
     "HTTP 200 — id coincide.","Pasa"),
    ("CP-093","Pedidos","Obtener pedido inexistente devuelve 404",
     "DB vacía","GET /orders/9999",
     "HTTP 404.","Pasa"),
    ("CP-094","Pedidos","Agregar ítem a pedido pendiente",
     "Pedido pending, producto con stock=100","POST /orders/{id}/items {product_id, qty:'3.000', price:'10.00'}",
     "HTTP 200 — items tiene 1 elemento, total=30.0.","Pasa"),
    ("CP-095","Pedidos","Agregar ítem con stock insuficiente devuelve 400",
     "Producto con stock=2","POST /orders/{id}/items {qty:'10.000'}",
     "HTTP 400 — detail contiene 'insufficient'.","Pasa"),
    ("CP-096","Pedidos","Agregar ítem de producto inactivo devuelve 400",
     "Producto inactivo (is_active=False)","POST /orders/{id}/items",
     "HTTP 400.","Pasa"),
    ("CP-097","Pedidos","Agregar ítem a pedido inexistente devuelve 400",
     "DB vacía","POST /orders/9999/items",
     "HTTP 400 — detail contiene 'not found'.","Pasa"),
    ("CP-098","Pedidos","Eliminar ítem de pedido pendiente",
     "Pedido con 1 ítem","DELETE /orders/{id}/items/{item_id}",
     "HTTP 200 — items=[].","Pasa"),
    ("CP-099","Pedidos","Eliminar pedido pendiente exitosamente",
     "Pedido en estado pending","DELETE /orders/{id}",
     "HTTP 204.","Pasa"),
    ("CP-100","Pedidos","Eliminar pedido no-pending devuelve 400",
     "Pedido en estado processing","DELETE /orders/{id}",
     "HTTP 400.","Pasa"),
    ("CP-101","Pedidos","Avanzar pending→processing descuenta stock",
     "Pedido con 1 ítem qty=5, producto stock=100",
     "POST /orders/{id}/advance",
     "HTTP 200 — status='processing'; stock del producto pasa a 95.","Pasa"),
    ("CP-102","Pedidos","Avanzar pending→processing crea movimiento de salida",
     "Pedido con ítem qty=3","POST /orders/{id}/advance",
     "StockMovement de tipo 'exit' con quantity=3 y order_id correcto.","Pasa"),
    ("CP-103","Pedidos","Ciclo completo: pending→processing→shipped→delivered",
     "Pedido con 1 ítem, stock suficiente",
     "3 × POST /orders/{id}/advance",
     "Statuses secuenciales: ['processing','shipped','delivered'].","Pasa"),
    ("CP-104","Pedidos","Avanzar pedido ya entregado devuelve 400",
     "Pedido en estado delivered","POST /orders/{id}/advance",
     "HTTP 400.","Pasa"),
    ("CP-105","Pedidos","Avanzar pedido vacío devuelve 400",
     "Pedido sin ítems","POST /orders/{id}/advance",
     "HTTP 400 — detail contiene 'no items'.","Pasa"),
    ("CP-106","Pedidos","Avanzar con stock insuficiente al momento de procesar devuelve 400",
     "Pedido con ítem qty=5, stock drenado a 0 manualmente",
     "POST /orders/{id}/advance",
     "HTTP 400. El pedido permanece en 'pending'.","Pasa"),

    # ═══ MOVIMIENTOS DE STOCK ═════════════════════════════════════════════════
    ("CP-107","Movimientos de Stock","Listar movimientos — vacío sin stock inicial",
     "Producto creado con stock=0","GET /stock-movements",
     "HTTP 200 — devuelve [].","Pasa"),
    ("CP-108","Movimientos de Stock","Producto con stock>0 genera movimiento de entrada al crearse",
     "Producto con stock=50","GET /stock-movements",
     "1 movimiento de tipo 'entry'.","Pasa"),
    ("CP-109","Movimientos de Stock","Filtro por tipo=entry",
     "1 movimiento de entrada","GET /stock-movements?type=entry",
     "Todos los resultados tienen type='entry'.","Pasa"),
    ("CP-110","Movimientos de Stock","Filtro por tipo=exit — sin ventas devuelve vacío",
     "Solo hay movimientos de entrada","GET /stock-movements?type=exit",
     "Devuelve [].","Pasa"),
    ("CP-111","Movimientos de Stock","Filtro por tipo=adjustment tras edición manual de stock",
     "Producto actualizado manualmente (stock 50→60)","GET /stock-movements?type=adjustment",
     "1 movimiento de tipo 'adjustment'.","Pasa"),
    ("CP-112","Movimientos de Stock","Filtro por product_id filtra correctamente",
     "2 productos con movimientos","GET /stock-movements?product_id={p1.id}",
     "Todos los movimientos pertenecen al producto p1.","Pasa"),
    ("CP-113","Movimientos de Stock","Filtro por rango de fechas",
     "1 movimiento existente","GET /stock-movements?date_from=2020-01-01&date_to=2099-12-31",
     "HTTP 200 — al menos 1 resultado.","Pasa"),
    ("CP-114","Movimientos de Stock","Paginación con limit=3",
     "5 productos creados (5 movimientos)","GET /stock-movements?limit=3",
     "HTTP 200 — array de 3 o menos elementos.","Pasa"),
    ("CP-115","Movimientos de Stock","Listar movimientos requiere autenticación",
     "Sin token","GET /stock-movements",
     "HTTP 401.","Pasa"),
    ("CP-116","Movimientos de Stock","Obtener movimiento por ID existente",
     "Movimiento en DB","GET /stock-movements/{id}",
     "HTTP 200 — id coincide, type='entry'.","Pasa"),
    ("CP-117","Movimientos de Stock","Obtener movimiento inexistente devuelve 404",
     "DB vacía","GET /stock-movements/9999",
     "HTTP 404.","Pasa"),
    ("CP-118","Movimientos de Stock","Movimiento de salida queda vinculado al pedido",
     "Pedido avanzado a 'processing'","GET /stock-movements/{exit_movement_id}",
     "HTTP 200 — order_id coincide con el pedido.","Pasa"),
]

# ── Excel builder ──────────────────────────────────────────────────────────────
HEADERS = [
    ("ID",                    8),
    ("Módulo",               16),
    ("Nombre del caso",      40),
    ("Precondiciones",       28),
    ("Entrada / Acción",     36),
    ("Resultado esperado",   42),
    ("Estado",                8),
]


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Casos de Prueba"

    # Title
    ws.merge_cells("A1:G1")
    tc = ws["A1"]
    tc.value = "StockFlow CRM — Casos de Prueba (Backend)"
    tc.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    tc.fill = _fill(C_HEADER)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    sc = ws["A2"]
    sc.value = (
        "Suite de tests: pytest + httpx + pytest-mock | "
        "Base de datos: SQLite in-memory | "
        "Total: 118 casos | Resultado esperado: 0 fallos"
    )
    sc.font = Font(name="Calibri", size=9, italic=True, color="374151")
    sc.fill = _fill("EFF6FF")
    sc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Column headers
    for ci, (hdr, w) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=ci, value=hdr)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = _fill("374151")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"

    current_module = None
    row = 4

    for tc_data in TEST_CASES:
        cp_id, module, name, precond, action, expected, status = tc_data

        # Module separator
        if module != current_module:
            current_module = module
            ws.merge_cells(f"A{row}:G{row}")
            mc = ws[f"A{row}"]
            mc.value = f"  MÓDULO: {module}"
            mc.font = Font(name="Calibri", size=10, bold=True, color="111827")
            mc.fill = _fill(MODULE_COLORS.get(module, "F3F4F6"))
            mc.alignment = Alignment(vertical="center")
            mc.border = _border()
            ws.row_dimensions[row].height = 18
            row += 1

        # Status color
        if status == "Pasa":
            status_fill = _fill(C_PASS)
        elif status == "Falla":
            status_fill = _fill(C_FAIL)
        else:
            status_fill = _fill(C_PEND)

        bg = C_ROW_A if row % 2 == 0 else C_ROW_B
        values = [cp_id, module, name, precond, action, expected, status]

        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = _border()
            cell.alignment = _wrap("center" if ci in (1, 7) else "left")
            if ci == 7:
                cell.fill = status_fill
                cell.font = Font(name="Calibri", size=9, bold=True,
                                 color="065F46" if status == "Pasa" else "991B1B")
            elif ci == 1:
                cell.fill = _fill(bg)
                cell.font = Font(name="Calibri", size=9, bold=True, color="1A56DB")
            else:
                cell.fill = _fill(bg)
                cell.font = _font(9)

        ws.row_dimensions[row].height = max(28, expected.count("\n") * 13 + 16)
        row += 1

    # Summary row
    pass_count = sum(1 for t in TEST_CASES if t[6] == "Pasa")
    fail_count = sum(1 for t in TEST_CASES if t[6] == "Falla")
    ws.merge_cells(f"A{row}:F{row}")
    sc = ws[f"A{row}"]
    sc.value = f"  RESUMEN:  {pass_count} casos Pasan  |  {fail_count} casos Fallan  |  Total: {len(TEST_CASES)}"
    sc.font = Font(name="Calibri", size=10, bold=True, color="111827")
    sc.fill = _fill("D1FAE5" if fail_count == 0 else "FEE2E2")
    sc.alignment = Alignment(vertical="center")
    sc.border = _border()
    ws.row_dimensions[row].height = 20

    output = "Casos_de_Prueba_StockFlow.xlsx"
    wb.save(output)
    print(f"Excel generado: {output}  ({pass_count}/{len(TEST_CASES)} pasan)")


if __name__ == "__main__":
    build()
