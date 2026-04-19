"""
Genera el documento de Casos de Uso de StockFlow CRM en formato Excel (.xlsx).
Ejecutar: python generate_use_cases.py
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── palette ────────────────────────────────────────────────────────────────────
C_HEADER  = "1A56DB"   # blue  — section headers
C_LABEL   = "374151"   # dark gray — field labels
C_ROW_A   = "EFF6FF"   # light blue — alternate rows
C_ROW_B   = "FFFFFF"   # white
C_BORDER  = "9CA3AF"   # gray borders
C_TITLE   = "111827"   # near-black
C_MOD     = "D1FAE5"   # green tint — module dividers


def thin_border(color=C_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def hdr_font(size=10, bold=True, color="FFFFFF"):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def body_font(size=9, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def wrap():
    return Alignment(wrap_text=True, vertical="top")


# ── use-case data ──────────────────────────────────────────────────────────────
USE_CASES = [
    # ── Autenticación ──────────────────────────────────────────────────────────
    {
        "id": "UC-01", "module": "Autenticación",
        "name": "Iniciar sesión",
        "actor": "Usuario (admin / operator)",
        "description": "El usuario ingresa sus credenciales para acceder al sistema y obtener un token JWT.",
        "preconditions": "El usuario tiene una cuenta registrada en el sistema.",
        "main_flow": (
            "1. El usuario navega a la pantalla de login.\n"
            "2. Ingresa su email y contraseña.\n"
            "3. Hace clic en «Iniciar sesión».\n"
            "4. El sistema valida las credenciales contra la base de datos.\n"
            "5. El sistema genera un token JWT con vigencia de 60 minutos.\n"
            "6. El sistema redirige al usuario al panel principal."
        ),
        "alt_flow": (
            "4a. Credenciales incorrectas: el sistema muestra «Email o contraseña inválidos» y no genera token.\n"
            "4b. Email no registrado: mismo mensaje de error (no se revela si el email existe)."
        ),
        "postconditions": "El usuario queda autenticado. El token JWT se almacena en el cliente y se incluye en cada petición posterior.",
    },
    {
        "id": "UC-02", "module": "Autenticación",
        "name": "Registrar usuario",
        "actor": "Administrador",
        "description": "Se crea una cuenta de usuario nueva en el sistema con un rol determinado.",
        "preconditions": "El email no debe estar registrado previamente.",
        "main_flow": (
            "1. El administrador realiza una petición POST a /auth/register con email, password y role.\n"
            "2. El sistema valida que el email tenga formato válido y la contraseña no esté vacía.\n"
            "3. El sistema hashea la contraseña con bcrypt.\n"
            "4. El sistema guarda el nuevo usuario en la base de datos.\n"
            "5. El sistema devuelve los datos del usuario creado (sin contraseña)."
        ),
        "alt_flow": (
            "2a. Email duplicado: el sistema devuelve HTTP 400 «Email already registered».\n"
            "2b. Email con formato inválido: el sistema devuelve HTTP 422."
        ),
        "postconditions": "El usuario existe en la base de datos y puede iniciar sesión.",
    },
    {
        "id": "UC-03", "module": "Autenticación",
        "name": "Consultar usuario actual",
        "actor": "Usuario autenticado",
        "description": "El sistema devuelve los datos del usuario asociado al token JWT activo.",
        "preconditions": "El usuario tiene un token JWT válido y no expirado.",
        "main_flow": (
            "1. El cliente realiza GET /auth/me incluyendo el token en el header Authorization.\n"
            "2. El sistema valida y decodifica el token.\n"
            "3. El sistema devuelve id, email y role del usuario."
        ),
        "alt_flow": (
            "2a. Token ausente o malformado: HTTP 401.\n"
            "2b. Token expirado: HTTP 401."
        ),
        "postconditions": "El cliente obtiene la información del usuario autenticado.",
    },

    # ── Productos ──────────────────────────────────────────────────────────────
    {
        "id": "UC-04", "module": "Productos",
        "name": "Crear producto",
        "actor": "Usuario autenticado",
        "description": "Se registra un nuevo producto en el catálogo de inventario.",
        "preconditions": "El usuario está autenticado. El SKU no existe en el sistema.",
        "main_flow": (
            "1. El usuario completa el formulario: SKU, nombre, precio, stock inicial y stock mínimo.\n"
            "2. El usuario envía el formulario.\n"
            "3. El sistema valida unicidad del SKU.\n"
            "4. El sistema crea el producto en la base de datos.\n"
            "5. Si el stock inicial > 0, el sistema registra automáticamente un movimiento de tipo «Entrada».\n"
            "6. El sistema devuelve el producto creado con HTTP 201."
        ),
        "alt_flow": (
            "3a. SKU duplicado: HTTP 400 «SKU already exists».\n"
            "5a. Stock inicial = 0: no se genera movimiento de entrada."
        ),
        "postconditions": "El producto aparece en el inventario. Si tenía stock inicial, existe un movimiento de entrada asociado.",
    },
    {
        "id": "UC-05", "module": "Productos",
        "name": "Actualizar producto",
        "actor": "Usuario autenticado",
        "description": "Se modifican los datos de un producto existente, incluyendo su stock.",
        "preconditions": "El producto existe en el sistema.",
        "main_flow": (
            "1. El usuario selecciona el producto y modifica los campos deseados (nombre, precio, stock, etc.).\n"
            "2. El usuario guarda los cambios.\n"
            "3. El sistema actualiza los datos en la base de datos.\n"
            "4. Si el campo «stock actual» cambió, el sistema calcula la diferencia y registra un movimiento de tipo «Ajuste».\n"
            "5. El sistema devuelve el producto actualizado."
        ),
        "alt_flow": (
            "1a. Producto no encontrado: HTTP 404.\n"
            "4a. Stock no cambió: no se genera movimiento."
        ),
        "postconditions": "El producto queda actualizado. Si el stock cambió, hay un movimiento de ajuste en el historial.",
    },
    {
        "id": "UC-06", "module": "Productos",
        "name": "Eliminar producto",
        "actor": "Usuario autenticado",
        "description": "Se elimina un producto del catálogo.",
        "preconditions": "El producto existe y no tiene movimientos de stock ni ítems de pedido asociados.",
        "main_flow": (
            "1. El usuario selecciona el producto y confirma la eliminación.\n"
            "2. El sistema verifica que no existan registros relacionados.\n"
            "3. El sistema elimina el producto.\n"
            "4. El sistema devuelve HTTP 204."
        ),
        "alt_flow": (
            "2a. El producto tiene movimientos de stock asociados: HTTP 409 «Cannot delete product with stock history».\n"
            "1a. Producto no encontrado: HTTP 404."
        ),
        "postconditions": "El producto ya no existe en el sistema.",
    },

    # ── Proveedores ────────────────────────────────────────────────────────────
    {
        "id": "UC-07", "module": "Proveedores",
        "name": "Crear proveedor",
        "actor": "Usuario autenticado",
        "description": "Se registra un nuevo proveedor en el sistema.",
        "preconditions": "El usuario está autenticado.",
        "main_flow": (
            "1. El usuario completa el formulario: nombre (obligatorio), contacto (obligatorio), email (obligatorio), teléfono (opcional).\n"
            "2. El usuario envía el formulario.\n"
            "3. El sistema guarda el proveedor en la base de datos.\n"
            "4. El sistema devuelve el proveedor creado con HTTP 201."
        ),
        "alt_flow": "2a. Campos obligatorios faltantes: HTTP 422.",
        "postconditions": "El proveedor queda disponible para asociar a facturas.",
    },
    {
        "id": "UC-08", "module": "Proveedores",
        "name": "Actualizar / Eliminar proveedor",
        "actor": "Usuario autenticado",
        "description": "Se modifican o eliminan los datos de un proveedor existente.",
        "preconditions": "El proveedor existe.",
        "main_flow": (
            "Actualizar:\n"
            "1. El usuario edita los campos del proveedor y guarda.\n"
            "2. El sistema actualiza los datos y devuelve HTTP 200.\n\n"
            "Eliminar:\n"
            "1. El usuario confirma la eliminación.\n"
            "2. El sistema elimina el proveedor y devuelve HTTP 204."
        ),
        "alt_flow": "Proveedor no encontrado: HTTP 404 en ambos casos.",
        "postconditions": "Los datos del proveedor quedan actualizados o el proveedor deja de existir.",
    },

    # ── Facturas ───────────────────────────────────────────────────────────────
    {
        "id": "UC-09", "module": "Facturas",
        "name": "Procesar factura con IA",
        "actor": "Usuario autenticado",
        "description": (
            "El usuario sube un archivo de factura (PDF/JPG/PNG) y el sistema lo envía a "
            "Gemini 2.5 Flash para extraer los ítems automáticamente."
        ),
        "preconditions": "El usuario está autenticado. La clave GOOGLE_API_KEY está configurada. El archivo pesa ≤ 20 MB.",
        "main_flow": (
            "1. El usuario selecciona el archivo de factura.\n"
            "2. El usuario hace clic en «Analizar con IA».\n"
            "3. El sistema valida el tipo MIME (PDF, JPG, PNG).\n"
            "4. El sistema envía el archivo a la API de Gemini 2.5 Flash.\n"
            "5. Gemini devuelve un JSON con los ítems detectados (descripción, cantidad, precio, confianza).\n"
            "6. El sistema crea una factura en estado «Pendiente» y guarda los ítems.\n"
            "7. El sistema devuelve los ítems con sugerencias de productos del catálogo.\n"
            "8. El sistema responde HTTP 201."
        ),
        "alt_flow": (
            "3a. Tipo MIME inválido: HTTP 415.\n"
            "5a. Gemini devuelve respuesta no parseable (JSON inválido): HTTP 422.\n"
            "5b. Servicio de Gemini saturado (free trial): el sistema puede mostrar error temporario; se recomienda reintentar en 10-30 segundos."
        ),
        "postconditions": "Se crea una factura en estado «Pendiente» con sus ítems en la base de datos.",
    },
    {
        "id": "UC-10", "module": "Facturas",
        "name": "Confirmar factura",
        "actor": "Usuario autenticado",
        "description": "El usuario revisa los ítems detectados, los asigna a productos del catálogo y confirma la factura, actualizando el stock.",
        "preconditions": "La factura existe y está en estado «Pendiente».",
        "main_flow": (
            "1. El usuario revisa cada ítem de la factura.\n"
            "2. Para cada ítem el usuario asigna un producto existente o crea uno nuevo.\n"
            "3. Opcionalmente ingresa el SKU del proveedor para cada ítem.\n"
            "4. El usuario puede marcar ítems como «Omitir» para no actualizar su stock.\n"
            "5. El usuario hace clic en «Confirmar factura».\n"
            "6. El sistema actualiza el stock de cada producto confirmado (suma la cantidad).\n"
            "7. El sistema registra movimientos de tipo «Entrada» por cada ítem procesado.\n"
            "8. El sistema guarda el mapeo SKU proveedor → producto para futuras facturas.\n"
            "9. El sistema actualiza el estado de la factura a «Confirmada»."
        ),
        "alt_flow": (
            "Pre. Factura no encontrada: HTTP 400.\n"
            "Pre. Factura ya confirmada: HTTP 400."
        ),
        "postconditions": "La factura está «Confirmada». El stock de los productos fue actualizado. Los movimientos de entrada están registrados.",
    },
    {
        "id": "UC-11", "module": "Facturas",
        "name": "Rechazar factura",
        "actor": "Usuario autenticado",
        "description": "El usuario rechaza una factura pendiente sin actualizar el stock.",
        "preconditions": "La factura existe y está en estado «Pendiente».",
        "main_flow": (
            "1. El usuario hace clic en «Rechazar».\n"
            "2. El sistema actualiza el estado de la factura a «Rechazada».\n"
            "3. El sistema devuelve la factura actualizada."
        ),
        "alt_flow": (
            "Pre. Factura no encontrada: HTTP 400.\n"
            "Pre. Factura ya rechazada: HTTP 400."
        ),
        "postconditions": "La factura queda en estado «Rechazada». El stock no fue modificado.",
    },

    # ── Clientes ───────────────────────────────────────────────────────────────
    {
        "id": "UC-12", "module": "Clientes",
        "name": "Crear / Actualizar / Eliminar cliente",
        "actor": "Usuario autenticado",
        "description": "CRUD completo de clientes del negocio.",
        "preconditions": "El usuario está autenticado.",
        "main_flow": (
            "Crear:\n"
            "1. El usuario completa nombre, email, teléfono y (opcionalmente) dirección.\n"
            "2. El sistema guarda el cliente y devuelve HTTP 201.\n\n"
            "Actualizar:\n"
            "1. El usuario edita los campos y guarda.\n"
            "2. El sistema actualiza y devuelve HTTP 200.\n\n"
            "Eliminar:\n"
            "1. El usuario confirma la eliminación.\n"
            "2. El sistema elimina el cliente y devuelve HTTP 204."
        ),
        "alt_flow": (
            "Crear — email inválido: HTTP 422.\n"
            "Actualizar / Eliminar — cliente no encontrado: HTTP 404."
        ),
        "postconditions": "El cliente existe/está actualizado/fue eliminado del sistema.",
    },
    {
        "id": "UC-13", "module": "Clientes",
        "name": "Ver historial de pedidos del cliente",
        "actor": "Usuario autenticado",
        "description": "Se consulta el historial completo de pedidos de un cliente con totales.",
        "preconditions": "El cliente existe.",
        "main_flow": (
            "1. El usuario hace clic en el ícono de historial del cliente.\n"
            "2. El sistema devuelve la lista de pedidos del cliente con estado, fecha y total.\n"
            "3. Si el cliente no tiene pedidos, se devuelve una lista vacía."
        ),
        "alt_flow": "Cliente no encontrado: HTTP 404.",
        "postconditions": "El usuario puede ver el historial comercial del cliente.",
    },

    # ── Pedidos ────────────────────────────────────────────────────────────────
    {
        "id": "UC-14", "module": "Pedidos",
        "name": "Crear pedido",
        "actor": "Usuario autenticado",
        "description": "Se crea un pedido vacío en estado «Pendiente» para un cliente.",
        "preconditions": "El cliente existe en el sistema.",
        "main_flow": (
            "1. El usuario selecciona el cliente en el formulario.\n"
            "2. El usuario hace clic en «Crear».\n"
            "3. El sistema crea el pedido en estado «Pendiente» sin ítems.\n"
            "4. El sistema devuelve el pedido con HTTP 201."
        ),
        "alt_flow": "Cliente no encontrado: HTTP 400.",
        "postconditions": "Existe un pedido vacío en estado «Pendiente» para el cliente.",
    },
    {
        "id": "UC-15", "module": "Pedidos",
        "name": "Agregar ítem al pedido",
        "actor": "Usuario autenticado",
        "description": "Se añade un producto con cantidad y precio unitario a un pedido pendiente.",
        "preconditions": "El pedido existe en estado «Pendiente». El producto está activo.",
        "main_flow": (
            "1. El usuario selecciona el producto, ingresa la cantidad y el precio unitario.\n"
            "2. El usuario hace clic en «Agregar».\n"
            "3. El sistema valida que el stock disponible sea ≥ cantidad solicitada.\n"
            "4. El sistema agrega el ítem al pedido.\n"
            "5. El sistema devuelve el pedido actualizado con el nuevo total."
        ),
        "alt_flow": (
            "3a. Stock insuficiente: HTTP 400 «Insufficient stock».\n"
            "3b. Producto inactivo: HTTP 400.\n"
            "Pre. Pedido no encontrado: HTTP 400."
        ),
        "postconditions": "El ítem aparece en el pedido. El total del pedido se actualizó.",
    },
    {
        "id": "UC-16", "module": "Pedidos",
        "name": "Avanzar estado del pedido",
        "actor": "Usuario autenticado",
        "description": "El pedido avanza un paso en el flujo: Pendiente → Procesando → Enviado → Entregado.",
        "preconditions": "El pedido existe. El pedido no está en estado «Entregado».",
        "main_flow": (
            "1. El usuario hace clic en «Avanzar».\n"
            "Pendiente → Procesando:\n"
            "  a. El sistema verifica que el pedido tenga al menos un ítem.\n"
            "  b. El sistema verifica que el stock de cada producto sea suficiente.\n"
            "  c. El sistema descuenta el stock de cada producto.\n"
            "  d. El sistema registra movimientos de tipo «Salida» por cada ítem.\n"
            "  e. El sistema envía email al cliente (si SendGrid está configurado).\n"
            "Procesando → Enviado:\n"
            "  a. El sistema actualiza el estado. Envía email al cliente.\n"
            "Enviado → Entregado:\n"
            "  a. El sistema actualiza el estado. Envía email al cliente."
        ),
        "alt_flow": (
            "Pre. Pedido sin ítems al avanzar a Procesando: HTTP 400 «Order has no items».\n"
            "b. Stock insuficiente en el momento de procesar: HTTP 400. El pedido permanece en «Pendiente».\n"
            "Pre. Pedido ya «Entregado»: HTTP 400."
        ),
        "postconditions": "El pedido avanzó de estado. Si pasó a «Procesando», el stock fue descontado y hay movimientos de salida.",
    },
    {
        "id": "UC-17", "module": "Pedidos",
        "name": "Eliminar pedido",
        "actor": "Usuario autenticado",
        "description": "Se elimina un pedido que aún no fue procesado.",
        "preconditions": "El pedido existe y está en estado «Pendiente».",
        "main_flow": (
            "1. El usuario confirma la eliminación.\n"
            "2. El sistema elimina el pedido y sus ítems.\n"
            "3. El sistema devuelve HTTP 204."
        ),
        "alt_flow": "Pedido en estado ≠ «Pendiente»: HTTP 400.",
        "postconditions": "El pedido ya no existe. El stock no fue afectado.",
    },

    # ── Movimientos de Stock ───────────────────────────────────────────────────
    {
        "id": "UC-18", "module": "Movimientos de Stock",
        "name": "Consultar movimientos de stock",
        "actor": "Usuario autenticado",
        "description": "El usuario puede ver y filtrar el historial completo de movimientos de stock (solo lectura).",
        "preconditions": "El usuario está autenticado.",
        "main_flow": (
            "1. El usuario navega al módulo «Movimientos».\n"
            "2. El sistema devuelve la lista de movimientos ordenada por fecha descendente.\n"
            "3. El usuario puede aplicar filtros: tipo (entrada/salida/ajuste), producto y rango de fechas.\n"
            "4. El sistema devuelve los movimientos que coinciden con los filtros.\n"
            "5. El usuario puede ver el detalle de un movimiento (factura o pedido vinculado)."
        ),
        "alt_flow": "Movimiento no encontrado en detalle: HTTP 404.",
        "postconditions": "El usuario puede auditar el historial de stock.",
    },
    {
        "id": "UC-19", "module": "Movimientos de Stock",
        "name": "Filtrar y paginar movimientos",
        "actor": "Usuario autenticado",
        "description": "El usuario aplica filtros combinados para acotar el historial de movimientos.",
        "preconditions": "Existen movimientos de stock registrados.",
        "main_flow": (
            "1. El usuario selecciona filtros: tipo de movimiento, producto específico y/o rango de fechas.\n"
            "2. El usuario puede configurar el límite de resultados por página.\n"
            "3. El sistema devuelve los movimientos que cumplen todos los criterios."
        ),
        "alt_flow": "Sin resultados: se devuelve una lista vacía (HTTP 200).",
        "postconditions": "El usuario obtiene el subconjunto de movimientos solicitado.",
    },
]

# ── Excel builder ──────────────────────────────────────────────────────────────
COLS = [
    ("ID",             8),
    ("Módulo",        14),
    ("Nombre del UC", 28),
    ("Actor",         20),
    ("Descripción",   38),
    ("Precondiciones",30),
    ("Flujo principal",52),
    ("Flujo alternativo / Excepciones", 42),
    ("Postcondiciones",30),
]

MODULES = {
    "Autenticación":      "D1FAE5",
    "Productos":          "FEF9C3",
    "Proveedores":        "FCE7F3",
    "Facturas":           "E0F2FE",
    "Clientes":           "FEE2E2",
    "Pedidos":            "EDE9FE",
    "Movimientos de Stock": "F3F4F6",
}


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Casos de Uso"

    # ── title row ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    tc = ws["A1"]
    tc.value = "StockFlow CRM — Casos de Uso"
    tc.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    tc.fill = fill(C_HEADER)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    sc = ws["A2"]
    sc.value = "Sistema: Mini CRM para negocios de ecommerce  |  Proyecto Final de Grado — Técnico en Programación"
    sc.font = Font(name="Calibri", size=9, italic=True, color="374151")
    sc.fill = fill("EFF6FF")
    sc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── column headers ─────────────────────────────────────────────────────────
    for col_idx, (header, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = hdr_font(10)
        cell.fill = fill(C_LABEL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 22

    # ── freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── data rows ──────────────────────────────────────────────────────────────
    current_module = None
    row_num = 4

    for uc in USE_CASES:
        # Module separator row
        if uc["module"] != current_module:
            current_module = uc["module"]
            ws.merge_cells(f"A{row_num}:I{row_num}")
            mc = ws[f"A{row_num}"]
            mc.value = f"  MÓDULO: {current_module}"
            mc.font = Font(name="Calibri", size=10, bold=True, color=C_TITLE)
            mc.fill = fill(MODULES.get(current_module, "F3F4F6"))
            mc.alignment = Alignment(vertical="center")
            mc.border = thin_border()
            ws.row_dimensions[row_num].height = 18
            row_num += 1

        # UC row
        values = [
            uc["id"],
            uc["module"],
            uc["name"],
            uc["actor"],
            uc["description"],
            uc["preconditions"],
            uc["main_flow"],
            uc["alt_flow"],
            uc["postconditions"],
        ]
        bg = C_ROW_A if row_num % 2 == 0 else C_ROW_B
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = body_font(9, bold=(col_idx == 1))
            cell.fill = fill(bg)
            cell.alignment = wrap()
            cell.border = thin_border()

        # Row height — taller for multi-line flows
        lines = uc["main_flow"].count("\n") + 1
        ws.row_dimensions[row_num].height = max(60, lines * 14)
        row_num += 1

    # ── save ──────────────────────────────────────────────────────────────────
    output = "Casos_de_Uso_StockFlow.xlsx"
    wb.save(output)
    print(f"Excel generado: {output}")


if __name__ == "__main__":
    build()
