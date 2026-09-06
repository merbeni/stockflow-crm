# -*- coding: utf-8 -*-
"""
Contrasta los números de la documentación contra la suite real.

Existe porque esos números se desfasan solos: cada vez que se agrega una prueba
hay que tocar el README, la tabla de totales del plan, el desglose por archivo y
la planilla de casos. Olvidarse de uno es lo normal, y una documentación que
dice 47 donde hay 55 deja de ser confiable para todo lo demás que afirma.

Uso:  python scripts/verificar_documentacion.py
Sale con código 1 si algo no coincide, así se puede enganchar a un hook.
"""
import collections
import pathlib
import re
import subprocess
import sys

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

RAIZ = pathlib.Path(__file__).resolve().parent.parent
PYTHON = RAIZ / "backend/venv/Scripts/python.exe"
if not PYTHON.exists():
    PYTHON = pathlib.Path(sys.executable)


def buscar(texto: str, patron: str):
    encontrado = re.search(patron, texto)
    return int(encontrado.group(1)) if encontrado else None


def main() -> int:
    salida = subprocess.run(
        [str(PYTHON), "-m", "pytest", "--collect-only", "-q"],
        cwd=RAIZ / "backend", capture_output=True, text=True,
    ).stdout
    lineas = [l for l in salida.splitlines() if "::" in l]
    backend = len(lineas)
    if backend == 0:
        print("No se pudo recolectar la suite de backend.")
        return 1

    archivos_front = sorted((RAIZ / "frontend/src").rglob("*.test.js*"))
    contar = lambda f: len(
        re.findall(r"^\s*(?:it|test)\(", f.read_text(encoding="utf-8"), re.M)
    )
    frontend = sum(contar(f) for f in archivos_front)

    from openpyxl import load_workbook
    hoja = load_workbook(RAIZ / "docs/test-cases/Casos_de_Prueba_StockFlow.xlsx").active
    planilla = sum(
        1 for r in range(4, hoja.max_row + 1)
        if str(hoja.cell(r, 1).value or "").startswith("CP-")
    )

    plan = (RAIZ / "docs/test-plan/PLAN_DE_PRUEBAS.md").read_text(encoding="utf-8")
    readme = (RAIZ / "README.md").read_text(encoding="utf-8")

    controles = [
        ("plan / backend", buscar(plan, r"\| Backend \(pytest\) \| \d+ \| (\d+) \|"), backend),
        ("plan / frontend", buscar(plan, r"\| Frontend \(Vitest\) \| \d+ \| (\d+) \|"), frontend),
        ("plan / archivos de frontend", buscar(plan, r"\| Frontend \(Vitest\) \| (\d+) \|"), len(archivos_front)),
        ("plan / total", buscar(plan, r"\| \*\*Total\*\* \| \*\*\d+\*\* \| \*\*(\d+)\*\* \|"), backend + frontend),
        ("plan / pirámide", buscar(plan, r"← (\d+) pruebas automatizadas"), backend),
        ("README / backend", buscar(readme, r"\((\d+) de backend"), backend),
        ("README / frontend", buscar(readme, r"\+ (\d+) de frontend"), frontend),
        ("README / total", buscar(readme, r"Estado actual: (\d+) pruebas"), backend + frontend),
        ("README / planilla", buscar(readme, r"\| (\d+) casos que documentan"), planilla),
    ]

    por_archivo = collections.Counter(l.split("::")[0].split("/")[-1] for l in lineas)
    for nombre, real in sorted(por_archivo.items()):
        controles.append((
            f"plan / {nombre}",
            buscar(plan, r"\| `" + re.escape(nombre) + r"` \| (\d+) \|"),
            real,
        ))
    for archivo in archivos_front:
        controles.append((
            f"plan / {archivo.name}",
            buscar(plan, r"\| `[^`]*" + re.escape(archivo.name) + r"` \| (\d+) \|"),
            contar(archivo),
        ))

    malas = 0
    for nombre, documenta, real in controles:
        if documenta != real:
            malas += 1
            print(f"  [MAL] {nombre:34} documenta={documenta}  real={real}")

    print(f"\nbackend {backend} + frontend {frontend} = {backend + frontend}"
          f"  |  planilla {planilla} casos")
    if malas:
        print(f"{malas} discrepancia(s): actualizá la documentación.")
        return 1
    print(f"Los {len(controles)} números documentados coinciden con la suite.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
